import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "서비스 활성화 지표"

# 색상
HEADER_BG  = "2B457A"
HEADER_FG  = "FFFFFF"
PROD_BG    = "D9E1F2"
INFER_BG   = "FFFACD"   # 노란색: 패턴 추론(*)
CONFIRM_BG = "FFE4E1"   # 연분홍: 확인 필요
FIX_BG     = "E8F5E9"   # 연녹색: 오타 수정 비고

thin = Side(style='thin', color="AAAAAA")
bd   = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell_fill(color):
    return PatternFill(fill_type='solid', fgColor=color)

def write_cell(ws, row, col, value, bold=False, bg=None, align_h='left', wrap=True, size=9):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, name='맑은 고딕', size=size,
                  color=HEADER_FG if bg == HEADER_BG else "000000")
    if bg:
        c.fill = cell_fill(bg)
    c.alignment = Alignment(horizontal=align_h, vertical='center', wrap_text=wrap)
    c.border = bd
    return c

# ── 헤더 ──────────────────────────────────────────────────────────────────────
headers = ["상품", "기준", "지표 명", "의미", "원천 데이터", "측정 주기",
           "데이터 대상 기간", "고객사 코드", "과금 기준", "운영 시스템",
           "시스템 운영자", "담당자", "소요 기간", "비고"]
ws.row_dimensions[1].height = 28
for ci, h in enumerate(headers, 1):
    write_cell(ws, 1, ci, h, bold=True, bg=HEADER_BG, align_h='center')

# ── 데이터 ────────────────────────────────────────────────────────────────────
# 각 행: [상품, 기준, 지표명, 의미, 원천데이터, 측정주기, 기간,
#         고객사코드, 과금기준, 운영시스템, 시스템운영자, 담당자, 소요기간, 비고]
rows = [
  # Knox Portal * Mail
  ["Knox Portal\n* Mail","User","MAU",
   "1개월 동안 1회 이상 PC 로그인 날짜가 있는 사용자 수\n(모바일은 상시 로그인)",
   "로그인 이벤트 로그 테이블","매월","과거 1년 ('25년 이후)",
   "확인 필요","등록 사용자(ID) 수","Knox Analytics","김혜경 프로님",
   "최동현/오진아/오세웅(Teams) 프로님","확인 필요",""],

  ["","User","등록 사용자 수",
   "서비스에 회원가입을 완료하여 ID가 생성된 계정의 누적 수",
   "회원가입 테이블 (가입일, 정상/탈퇴/휴면 Status)","매월","과거 1년 ('25년 이후)",
   "확인 필요","등록 사용자(ID) 수 *","Knox Analytics *","김혜경 프로님 *",
   "최동현/오진아/오세웅(Teams) 프로님 *","확인 필요",""],

  ["","Usage","Portal 메일 발신량","모바일+PC 포함",
   "메일 발신 이벤트 로그 테이블 *","매월 *","과거 1년 ('25년 이후) *",
   "확인 필요","확인 필요","Knox Analytics *","김혜경 프로님 *",
   "최동현/오진아/오세웅(Teams) 프로님 *","확인 필요",""],

  ["","Usage","Portal 메일 수신량","모바일+PC 포함",
   "메일 수신 이벤트 로그 테이블 *","매월 *","과거 1년 ('25년 이후) *",
   "확인 필요","확인 필요","Knox Analytics *","김혜경 프로님 *",
   "최동현/오진아/오세웅(Teams) 프로님 *","확인 필요",""],

  # Knox Teams * Messenger
  ["Knox Teams\n* Messenger","User","MAU",
   "1개월 동안 1회 이상 PC 로그인 날짜가 있는 사용자 수\n(모바일은 상시 로그인)",
   "로그인 이벤트 로그 테이블 *","매월 *","과거 1년 ('25년 이후) *",
   "확인 필요","등록 사용자(ID) 수 *","Knox Analytics *","김혜경 프로님 *",
   "최동현/오진아/오세웅(Teams) 프로님 *","확인 필요","=Portal과 동일"],

  ["","User","등록 사용자 수",
   "서비스에 회원가입을 완료하여 ID가 생성된 계정의 누적 수",
   "회원가입 테이블 (가입일, 정상/탈퇴/휴면 Status) *","매월 *","과거 1년 ('25년 이후) *",
   "확인 필요","등록 사용자(ID) 수 *","Knox Analytics *","김혜경 프로님 *",
   "최동현/오진아/오세웅(Teams) 프로님 *","확인 필요","=Portal과 동일"],

  ["","Usage","메시지 발신량","모바일+PC 포함",
   "메시지 발신 이벤트 로그 테이블 *","매월 *","과거 1년 ('25년 이후) *",
   "확인 필요","확인 필요","Knox Analytics *","김혜경 프로님 *",
   "최동현/오진아/오세웅(Teams) 프로님 *","확인 필요",""],

  ["","Usage","메시지 수신량 ✏","모바일+PC 포함",
   "메시지 수신 이벤트 로그 테이블 *","매월 *","과거 1년 ('25년 이후) *",
   "확인 필요","확인 필요","Knox Analytics *","김혜경 프로님 *",
   "최동현/오진아/오세웅(Teams) 프로님 *","확인 필요","오타 수정: 메신지→메시지"],

  # Knox Meeting
  ["Knox Meeting","User","MAU",
   "1개월 동안 1회 이상 로그인 날짜가 있는 사용자 수",
   "로그인 로그 테이블","매월","과거 3년 ('23년 이후)",
   "확인 필요","동접 최대 포트수, 사용자수","Knox Meeting Admin",
   "상품: 전기구/조영우 프로\n개발: 송영용/정희철 프로님\n운영: SRE그룹 장세찬/김인규 프로님",
   "최동현/전기구 프로님","확인 필요",""],

  ["","User","등록 사용자 수",
   "서비스에 회원가입을 완료하여 ID가 생성된 계정의 누적 수",
   "회원가입 테이블 (가입일, 정상/탈퇴/휴면 Status)","매월","과거 3년 ('23년 이후)",
   "확인 필요","동접 최대 포트수, 사용자수 *","Knox Meeting Admin *",
   "상품: 전기구/조영우 프로\n개발: 송영용/정희철 프로님\n운영: SRE그룹 장세찬/김인규 프로님 *",
   "최동현/전기구 프로님 *","확인 필요",""],

  ["","Usage","총 사용 시간\n총 회의건 수\n총 회의실 개소수",
   "미팅 총 사용 현황 (WC/VC/AC 구분)",
   "미팅 접속시간, 회의 건수(WC/VC/AC), 회의실 개소수","매월","과거 3년 ('23년 이후)",
   "확인 필요","확인 필요","Knox Meeting Admin *",
   "상품: 전기구/조영우 프로\n개발: 송영용/정희철 프로님\n운영: SRE그룹 장세찬/김인규 프로님 *",
   "최동현/전기구 프로님 *","확인 필요",""],

  # Knox Drive
  ["Knox Drive","Usage","스토리지 사용량","저장된 데이터 총 용량",
   "파일 메타데이터 / 스토리지 / Usage 데이터","매월","과거 3년 ('23년 이후)",
   "100개 미만","스토리지 사용량, 등록 사용자 수","Drive Admin Portal",
   "Knox Drive 운영팀(Drive개발그룹)\ndrive.tech@samsung.com",
   "신명호/김한비/김수정 프로님","3주",""],

  ["","User","MAU",
   "1개월 동안 1회 이상 로그인 날짜가 있는 사용자 수",
   "로그인 로그 테이블","매월","과거 3년 ('23년 이후)",
   "100개 미만 *","스토리지 사용량, 등록 사용자 수 *","Drive Admin Portal *",
   "Knox Drive 운영팀(Drive개발그룹)\ndrive.tech@samsung.com *",
   "신명호/김한비/김수정 프로님 *","확인 필요",""],

  ["","User","등록 사용자 수",
   "서비스에 회원가입을 완료하여 ID가 생성된 계정의 누적 수",
   "회원가입 테이블 (가입일, 정상/탈퇴/휴면 Status)","매월","과거 3년 ('23년 이후)",
   "100개 미만 *","스토리지 사용량, 등록 사용자 수 *","Drive Admin Portal *",
   "Knox Drive 운영팀(Drive개발그룹)\ndrive.tech@samsung.com *",
   "신명호/김한비/김수정 프로님 *","확인 필요",""],

  # Knox EFSS
  ["Knox EFSS","Usage","스토리지 사용량","저장된 데이터 총 용량",
   "파일 메타데이터 / 스토리지 / Usage 데이터","매월 *","과거 3년 ('23년 이후)",
   "확인 필요","스토리지 사용량, 등록 사용자 수","EFSS Admin Portal",
   "반도체: 김봉규 프로님\n디스플레이: 김경훈 프로님\n공용망: 엄재호 프로님\n운영망 셀장: 김영진 프로님",
   "김성덕 프로님","확인 필요","측정 주기 패턴 추론 *"],

  ["","User","MAU",
   "1개월 동안 1회 이상 로그인 날짜가 있는 사용자 수",
   "로그인 로그 테이블","매월","과거 3년 ('23년 이후)",
   "확인 필요","스토리지 사용량, 등록 사용자 수 *","EFSS Admin Portal *",
   "반도체: 김봉규 프로님\n디스플레이: 김경훈 프로님\n공용망: 엄재호 프로님\n운영망 셀장: 김영진 프로님 *",
   "김성덕 프로님 *","확인 필요",""],

  ["","User","등록 사용자 수",
   "서비스에 회원가입을 완료하여 ID가 생성된 계정의 누적 수",
   "회원가입 테이블 (가입일, 정상/탈퇴/휴면 Status)","매월","과거 3년 ('23년 이후)",
   "확인 필요","스토리지 사용량, 등록 사용자 수 *","EFSS Admin Portal *",
   "반도체: 김봉규 프로님\n디스플레이: 김경훈 프로님\n공용망: 엄재호 프로님\n운영망 셀장: 김영진 프로님 *",
   "김성덕 프로님 *","확인 필요",""],

  # Copilot
  ["Copilot\n(Knox Suite 포함)","Usage","기능 이용 횟수\n(프롬프트 개수)",
   "Copilot 기능 호출 총 횟수",
   "Knox Analytics 프롬프트 이벤트 로그 테이블 *","매월","과거 3년 ('23년 이후)",
   "확인 필요","확인 필요","Knox Analytics","확인 필요",
   "최동현 프로님","확인 필요",""],

  ["","User","MAU",
   "1개월 동안 Copilot 기능 호출 사용자 수",
   "로그인 로그 테이블","매월","과거 3년 ('23년 이후)",
   "확인 필요","확인 필요 *","Knox Analytics *","확인 필요",
   "최동현 프로님 *","확인 필요",""],

  ["","User","권한자 수",
   "Copilot 기능 이용가능자 수",
   "회원가입 테이블 (가입일, 정상/탈퇴/휴면 Status)","매월","과거 3년 ('23년 이후)",
   "확인 필요","확인 필요 *","Knox Analytics *","확인 필요",
   "최동현 프로님 *","확인 필요",""],

  # Brity Works
  ["Brity Works","확인 필요","복합 지표\n(Portal+Teams\n+Meeting+Drive)",
   "Knox Portal + Teams + Meeting + Drive 통합 지표",
   "각 서브 상품과 동일","매월 *","확인 필요",
   "확인 필요","확인 필요","각 서브 상품과 동일",
   "각 서브 상품과 동일 *","이명민/박건진 프로님","확인 필요",
   "복합 지표 - 서브 상품 지표 확정 후 정의"],

  # EMM Cloud / EMS / ZTM
  ["EMM Cloud","-","- (미정)","-","-","-","-","-","-","-","-","최재웅 프로님","-","지표 미정"],
  ["EMS",      "-","- (미정)","-","-","-","-","-","-","-","-","고동현 프로님","-","지표 미정"],
  ["ZTM",      "-","- (미정)","-","-","-","-","-","-","-","-","김동현 프로님","-","지표 미정"],
]

# ── 셀 작성 및 색상 적용 ────────────────────────────────────────────────────
merge_groups = [(2,5),(6,9),(10,12),(13,15),(16,18),(19,21),(22,22),(23,23),(24,24),(25,25)]

for ri, row_data in enumerate(rows):
    excel_row = ri + 2
    ws.row_dimensions[excel_row].height = 45

    for ci, val in enumerate(row_data):
        v = str(val) if val is not None else ""

        # 배경색 결정
        if ci == 0 and v:  # 상품명 셀
            bg = PROD_BG
        elif "확인 필요" in v:
            bg = CONFIRM_BG
        elif v.endswith(" *") or " * " in v or v.endswith("*"):
            bg = INFER_BG
        elif ci == 13 and "오타 수정" in v:
            bg = FIX_BG
        else:
            bg = None

        write_cell(ws, excel_row, ci+1, v,
                   bold=(ci == 0 and bool(v)),
                   bg=bg,
                   align_h='center' if ci in (0,1,5,6,12) else 'left')

# ── 상품 열 병합 ───────────────────────────────────────────────────────────
for (start, end) in merge_groups:
    if start < end:
        ws.merge_cells(start_row=start, start_column=1,
                       end_row=end,   end_column=1)
        c = ws.cell(row=start, column=1)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.fill = cell_fill(PROD_BG)
        c.font = Font(bold=True, name='맑은 고딕', size=9)
        c.border = bd

# ── 열 너비 ────────────────────────────────────────────────────────────────
col_widths = [14, 8, 18, 30, 28, 8, 16, 12, 18, 16, 28, 22, 10, 24]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── 범례 시트 ──────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("범례")
legends = [
    ("색상 범례", ""),
    ("배경색", "의미"),
    ("흰색 (기본)", "원본 데이터 그대로"),
    ("노란색", "같은 상품군 패턴으로 추론한 값 (*)"),
    ("연분홍", "담당자 확인 필요 항목"),
    ("연녹색", "오타 수정 내용"),
    ("연파랑 (상품명)", "상품 그룹 구분"),
    ("", ""),
    ("* 표기 의미", "동일 상품군 내 패턴 전파 또는 지표 유형별 추론 (원본 미기재)"),
    ("확인 필요", "담당자 확인 후 채워야 하는 항목"),
]
for ri, (k, v) in enumerate(legends, 1):
    ws2.cell(row=ri, column=1, value=k).font = Font(bold=(ri<=2), name='맑은 고딕', size=10)
    ws2.cell(row=ri, column=2, value=v).font = Font(name='맑은 고딕', size=10)
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 50

wb.save("/home/ubuntu/service_activation_indicators.xlsx")
print("Excel 저장 완료: /home/ubuntu/service_activation_indicators.xlsx")
