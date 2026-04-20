import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "효율수익_개발요청서"

# 색상 정의
NAVY = "1E3079"
NAVY_BG = "EEF2FB"
BLUE_BORDER = "89B2F7"
YELLOW_BG = "FFFBEA"
RED_BG = "FFF0F0"
HEADER_BG = "F0F4FB"
WHITE = "FFFFFF"
GRAY_ROW = "F8FAFD"
LIGHT_GRAY = "F5F5F5"

def navy_font(bold=False, size=11, color=NAVY):
    return Font(name="맑은 고딕", bold=bold, size=size, color=color)

def normal_font(bold=False, size=10, color="1A1A1A"):
    return Font(name="맑은 고딕", bold=bold, size=size, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_all(color="D0D8E8"):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def border_bottom(color="1E3079", width="medium"):
    return Border(bottom=Side(style=width, color=color))

def set_cell(ws, row, col, value, font=None, fill_color=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill_color: cell.fill = fill(fill_color)
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    return cell

def merge_and_set(ws, r1, c1, r2, c2, value, font=None, fill_color=None, alignment=None, border=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = set_cell(ws, r1, c1, value, font, fill_color, alignment, border)
    return cell

def section_title(ws, row, title, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name="맑은 고딕", bold=True, size=11, color=NAVY)
    cell.fill = fill(NAVY_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = Border(left=Side(style="thick", color=NAVY))
    ws.row_dimensions[row].height = 22
    return row + 1

def sub_title(ws, row, title, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name="맑은 고딕", bold=True, size=10, color="333333")
    cell.fill = fill("F0F4FB")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = Border(left=Side(style="medium", color=BLUE_BORDER))
    ws.row_dimensions[row].height = 20
    return row + 1

def note_row(ws, row, text, cols=6, bg=YELLOW_BG, color="7A6000"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="맑은 고딕", size=9, color=color, italic=True)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 18
    return row + 1

def tbd_row(ws, row, text, cols=6):
    return note_row(ws, row, text, cols, bg=RED_BG, color="990000")

def table_header(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start+i, value=h)
        cell.font = Font(name="맑은 고딕", bold=True, size=10, color=NAVY)
        cell.fill = fill(HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_all("D0D8E8")
    ws.row_dimensions[row].height = 20
    return row + 1

def table_row(ws, row, values, col_start=1, even=False):
    bg = GRAY_ROW if even else WHITE
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=col_start+i, value=v)
        cell.font = normal_font(size=10)
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border_all("E2E6EE")
    ws.row_dimensions[row].height = 18
    return row + 1

def blank(ws, row, h=6):
    ws.row_dimensions[row].height = h
    return row + 1

# ── 열 너비 설정
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 30
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 20
ws.column_dimensions["F"].width = 18

r = 1

# ══ 문서 헤더 ══
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
cell = ws.cell(row=r, column=1, value="[개발 요청] 사업건전성 대시보드 효율/수익 지표 산출 로직 및 화면 개발")
cell.font = Font(name="맑은 고딕", bold=True, size=14, color=NAVY)
cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[r].height = 30
r += 1

meta = [("요청 일자", "2026-04-20"), ("요청 부서", "솔루션사업부 기획팀"),
        ("대상 시스템", "사업건전성 대시보드"), ("문서 버전", "v0.3")]
for label, val in meta:
    ws.cell(row=r, column=1, value=label).font = Font(name="맑은 고딕", bold=True, size=9, color="555555")
    ws.cell(row=r, column=2, value=val).font = Font(name="맑은 고딕", size=9, color="1A1A1A")
    r += 1

# 헤더 하단 구분선
for col in range(1, 7):
    ws.cell(row=r-1, column=col).border = Border(bottom=Side(style="medium", color=NAVY))
r = blank(ws, r, 8)

# ══ 1. 요청 배경 ══
r = section_title(ws, r, "1. 요청 배경")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
cell = ws.cell(row=r, column=1,
    value="사업건전성 대시보드의 효율/수익 영역 지표 산출 로직을 드리니, 화면 개발을 요청합니다.")
cell.font = normal_font(size=10)
cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[r].height = 32
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
cell = ws.cell(row=r, column=1,
    value="원천 데이터(vf219, vf599 등)는 당사 DB에 이미 수신 중이며, Row 1~6 화면은 Tableau로 기 구현되어 있어 본 요청서의 개발 범위에서 제외합니다.")
cell.font = normal_font(size=10)
cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[r].height = 24
r += 1
r = blank(ws, r)

# ══ 2. 요청 사항 요약 ══
r = section_title(ws, r, "2. 요청 사항 요약")
r = table_header(ws, r, ["#", "구분", "내용", "우선순위", "납기 (목표)", ""])
rows2 = [
    ["1", "지표 산출 로직 개발", "LTV · CAC · LTV:CAC Ratio · CAC Payback 지표별 산출 기준에 따른 계산 로직 구현", "1차", "미정", ""],
    ["2", "화면", "Tableau 기 구현 — 본 요청 범위 제외 (와이어프레임 별도 제공 예정)", "-", "-", ""],
]
for i, row_data in enumerate(rows2):
    r = table_row(ws, r, row_data, even=(i%2==1))
r = note_row(ws, r, "※ 납기 일정은 협의 후 확정 예정입니다.")
r = blank(ws, r)

# ══ 3. 원천 데이터 현황 ══
r = section_title(ws, r, "3. 원천 데이터 현황")
r = note_row(ws, r, "아래 데이터는 당사 DB에 이미 수신 중이며, 별도 인터페이스 요청 없이 사용 가능합니다.", bg=WHITE, color="1A1A1A")
r = table_header(ws, r, ["#", "필드/테이블", "내용", "조회 조건", "히스토리 범위", "활용 지표"])
rows3 = [
    ["1", "vf219", "매출총계 (상품별 연도별)", "상품 코드 · 연도 기준 컬럼 — 개발팀 확인 필요", "2022년 ~ 현재", "LTV · ARPA · 이탈률"],
    ["2", "vf599", "영업/마케팅 비용 (상품별 연도별)", "상품 코드 · 연도 기준 컬럼 — 개발팀 확인 필요", "2022년 ~ 현재", "CAC"],
    ["3", "기타", "Revenue · Operating Profit · Plan · FTE · ARR 등", "개발팀 확인 필요", "Row 5 기준: 2023.10 ~ 현재", "Row 1·3·4·5·6 지표"],
]
for i, row_data in enumerate(rows3):
    r = table_row(ws, r, row_data, even=(i%2==1))
r = note_row(ws, r, "※ vf219는 RC/NRC 데이터 분리가 어려우므로, 실무 적용 시 전체 매출을 Recurring Revenue로 간주하여 사용합니다.")
r = blank(ws, r)

# ══ 4. 마스터 데이터 요청 ══
r = section_title(ws, r, "4. 마스터 데이터 요청 (화면 필터용)")
r = note_row(ws, r, "화면 상단 상품 드롭다운 필터 구현을 위해 아래 마스터 데이터가 필요합니다.", bg=WHITE, color="1A1A1A")
r = table_header(ws, r, ["#", "항목", "내용", "샘플값", "", ""])
rows4 = [
    ["1", "상품 목록", "필터 드롭다운에 표시될 상품명 목록 (코드 + 표시명)", "ZTM · Nexprime SCM · Knox Portal · EMS 등 13개", "", ""],
    ["2", "상품 그룹핑", "상품 개별 / 전체(All) 선택 구분", "Knox (All) = Knox Portal + Knox Meeting + Knox EFSS/Drive", "", ""],
    ["3", "EMS 예외 처리", "EMS / EMS(All) 선택 시 LTV:CAC 관련 카드 N/A 표시 처리용 플래그", "is_ltv_applicable: Y/N", "", ""],
]
for i, row_data in enumerate(rows4):
    r = table_row(ws, r, row_data, even=(i%2==1))
r = note_row(ws, r, "※ 상품 코드 체계 및 마스터 테이블명은 개발팀 확인 후 공유 바랍니다.")
r = blank(ws, r)

# ══ 5. 상품 분류 및 지표 산출 가능 여부 ══
r = section_title(ws, r, "5. 상품 분류 및 지표 산출 가능 여부")
r = table_header(ws, r, ["상품", "License Type", "Price Type", "RC/NRC", "LTV / CAC / LTV:CAC / CAC Payback", ""])
products = [
    ["ZTM", "Term", "Fixed", "RC", "✅ / ✅ / ✅ / ✅", ""],
    ["Nexprime SCM", "Term", "Fixed", "RC", "✅ / ✅ / ✅ / ✅", ""],
    ["Nexprime SCM Mobile", "Term", "Fixed", "RC", "✅ / ✅ / ✅ / ✅", ""],
    ["Nexprime SCM (All)", "Term", "Fixed", "RC", "✅ / ✅ / ✅ / ✅", ""],
    ["Nexprime HCM", "Term", "Fixed", "RC", "✅ / ✅ / ✅ / ✅", ""],
    ["Knox Portal", "Term", "Usage", "RC", "✅ / ✅ / ✅ / ✅", ""],
    ["Knox Meeting", "Term", "Usage", "RC", "✅ / ✅ / ✅ / ✅", ""],
    ["Knox EFSS/Drive", "Term", "Usage", "RC", "✅ / ✅ / ✅ / ✅", ""],
    ["Knox (All)", "Term", "Usage", "RC", "✅ / ✅ / ✅ / ✅", ""],
    ["EMM Cloud", "Term", "Usage", "RC", "✅ / ✅ / ✅ / ✅", ""],
    ["Brity Automation", "Term+Perpetual", "Fixed+One-time", "RC+Non-RC", "✅ / ✅ / ✅ / ✅", ""],
    ["EMS", "Term+Perpetual", "Fixed+One-time", "RC+Non-RC", "❌ / ✅ / ❌ / ❌", ""],
    ["EMS (All)", "Term+Perpetual", "Fixed+One-time", "RC+Non-RC", "❌ / ✅ / ❌ / ❌", ""],
]
for i, row_data in enumerate(products):
    r = table_row(ws, r, row_data, even=(i%2==1))
r = note_row(ws, r, "※ EMS / EMS (All): Recurring 비중 약 9% (라이선스+운영 149억 / 전체 1,572억), 데이터 분리 불가로 LTV · LTV:CAC · CAC Payback 산출 제외. CAC만 산출.")
r = blank(ws, r)

# ══ 6. 지표별 산출 기준 ══
r = section_title(ws, r, "6. 지표별 산출 기준")

indicators = [
    ("6-1. LTV (Customer Lifetime Value)", [
        ["산식 (이론)", "(Recurring 매출에 대한 ARPA ÷ 이탈률) + (연 일회성 매출 ÷ 고객 수)", "", "", "", ""],
        ["실무 적용 산식", "(sum(vf219) ÷ 고객 수 ÷ 이탈률) + (연 일회성 매출 ÷ 고객 수)", "", "", "", ""],
        ["비고", "RC/NRC 데이터 분리 불가로 vf219 전체를 Recurring으로 간주", "", "", "", ""],
        ["적용 상품", "EMS / EMS(All) 제외 전 상품", "", "", "", ""],
        ["표시 주기", "Yearly (2023 / 2024 / 2025)", "", "", "", ""],
        ["단위", "억원", "", "", "", ""],
    ]),
    ("  └ 구성 요소: ARPA / 이탈률", [
        ["ARPA 산식", "sum(vf219) ÷ 고객 수  (해당 연도 vf219 ≠ 0인 고객 수 기준)", "", "", "", ""],
        ["이탈률 산식", "(1 - (N-1년 sum(vf219) ÷ N-2년 sum(vf219))) × 100  ※ N-1년 ≥ N-2년이면 0 → LTV 빈칸", "", "", "", ""],
    ]),
    ("6-2. CAC (Customer Acquisition Cost)", [
        ["산식", "N-2년 sum(vf599) ÷ N-1년 신규 고객 수", "", "", "", ""],
        ["신규 고객 정의", "N-2년 vf219 = 0 AND N-1년 vf219 ≠ 0", "", "", "", ""],
        ["적용 상품", "전체 (EMS 포함)", "", "", "", ""],
        ["표시 주기", "Yearly (2023 / 2024 / 2025)", "", "", "", ""],
        ["단위", "억원", "", "", "", ""],
        ["예외", "신규 고객 수 = 0이면 빈칸", "", "", "", ""],
    ]),
    ("6-3. LTV:CAC Ratio", [
        ["산식", "LTV ÷ CAC", "", "", "", ""],
        ["적용 상품", "EMS / EMS(All) 제외 전 상품", "", "", "", ""],
        ["단위", "배수 (예: 1.6배)", "", "", "", ""],
        ["참고 기준", "LTV : CAC = 3 : 1이 적정 수준", "", "", "", ""],
        ["예외", "LTV 또는 CAC 빈칸이면 빈칸", "", "", "", ""],
    ]),
    ("6-4. CAC Payback Period", [
        ["산식", "(CAC ÷ ARPA) × 12", "", "", "", ""],
        ["적용 상품", "EMS / EMS(All) 제외 전 상품", "", "", "", ""],
        ["단위", "개월", "", "", "", ""],
        ["참고 기준", "B2B SaaS 적정 수준: 12개월 이하", "", "", "", ""],
        ["예외", "ARPA = 0 또는 CAC 빈칸이면 빈칸", "", "", "", ""],
    ]),
]

for title, rows in indicators:
    r = sub_title(ws, r, title)
    r = table_header(ws, r, ["항목", "내용", "", "", "", ""])
    for i, row_data in enumerate(rows):
        # merge columns 2~6 for value
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        cell_k = ws.cell(row=r, column=1, value=row_data[0])
        cell_v = ws.cell(row=r, column=2, value=row_data[1])
        bg = GRAY_ROW if i%2==1 else WHITE
        for col in [1, 2]:
            ws.cell(row=r, column=col).font = normal_font(size=10)
            ws.cell(row=r, column=col).fill = fill(bg)
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.cell(row=r, column=col).border = border_all("E2E6EE")
        ws.row_dimensions[r].height = 18
        r += 1

# 6-7 TBD
r = sub_title(ws, r, "6-7. Row 1 · 3 · 4 · 5 · 6 지표 (산식 확인 필요)")
r = tbd_row(ws, r, "※ 아래 지표는 사내 기존 BI 시스템에서 이미 산출 중입니다. 개발팀은 기존 산출 로직 및 데이터 출처(테이블/필드명)를 확인하여 동일 기준으로 적용하되, 확인된 산식을 기획팀에 공유 바랍니다.")
r = table_header(ws, r, ["Row", "지표명", "비고", "", "", ""])
tbd_rows = [
    ["Row 1", "Gross Sales Efficiency / Net Sales Efficiency", "기존 BI 로직 적용", "", "", ""],
    ["Row 3", "ARPA by ACV / ARPA by MRR (분기별)", "기존 BI 로직 적용", "", "", ""],
    ["Row 4", "Revenue on FTE / Rule of 40", "기존 BI 로직 적용 — FTE 데이터 출처 확인 필요", "", "", ""],
    ["Row 5", "Revenue & Operating Profit / % Gross Margin & Operating Profit", "기존 BI 로직 적용", "", "", ""],
    ["Row 6", "Revenue vs. Plan / Operating Profit vs. Plan", "기존 BI 로직 적용 — Plan 데이터 출처 확인 필요", "", "", ""],
]
for i, row_data in enumerate(tbd_rows):
    r = table_row(ws, r, row_data, even=(i%2==1))
r = blank(ws, r)

# ══ 7. 예외 처리 기준 ══
r = section_title(ws, r, "7. 예외 처리 기준 (공통)")
r = table_header(ws, r, ["조건", "처리 방식", "영향 지표", "", "", ""])
exc_rows = [
    ["이탈률 = 0", "빈칸 표시", "LTV · LTV:CAC · CAC Payback", "", "", ""],
    ["신규 고객 수 = 0", "빈칸 표시", "CAC · LTV:CAC · CAC Payback", "", "", ""],
    ["ARPA = 0", "빈칸 표시", "CAC Payback", "", "", ""],
    ["EMS / EMS(All) 선택 시", "LTV · LTV:CAC · CAC Payback 카드 N/A 표시", "Row 2 전체", "", "", ""],
]
for i, row_data in enumerate(exc_rows):
    r = table_row(ws, r, row_data, even=(i%2==1))
r = blank(ws, r)

# ══ 8. 화면 개발 스펙 ══
r = section_title(ws, r, "8. 화면 개발 스펙")
r = sub_title(ws, r, "8-1. 전체 레이아웃")
layout_rows = [
    ["구성", "6 Row × 2 카드 (총 12 카드)", "", "", "", ""],
    ["상단 필터", "상품 드롭다운 (전체 / 상품별 선택) — Section 4 마스터 데이터 기반", "", "", "", ""],
    ["표시 주기", "Yearly 기본 (Row 3·4·5·6은 분기/월별)", "", "", "", ""],
    ["페이지 배경", "#F5F5F5", "", "", "", ""],
    ["카드 스타일", "border-radius: 8px · box-shadow: 0 2px 8px rgba(0,0,0,0.08)", "", "", "", ""],
]
r = table_header(ws, r, ["항목", "내용", "", "", "", ""])
for i, row_data in enumerate(layout_rows):
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.cell(row=r, column=1, value=row_data[0]).font = normal_font(size=10)
    ws.cell(row=r, column=2, value=row_data[1]).font = normal_font(size=10)
    bg = GRAY_ROW if i%2==1 else WHITE
    for col in [1, 2]:
        ws.cell(row=r, column=col).fill = fill(bg)
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=r, column=col).border = border_all("E2E6EE")
    ws.row_dimensions[r].height = 18
    r += 1

r = sub_title(ws, r, "8-2. Row별 카드 스펙")
r = table_header(ws, r, ["Row", "카드 (좌)", "카드 (우)", "차트 타입", "X축", "단위"])
card_rows = [
    ["Row 1", "Gross Sales Efficiency", "Net Sales Efficiency", "단색 바 차트", "2022 / 2023 / 2024", "%"],
    ["Row 2", "CAC / LTV / LTV:CAC Ratio", "CAC Payback Period", "콤비네이션(좌) · 라인(우)", "2022 / 2023 / 2024", "억원 · 배수 / 개월"],
    ["Row 3", "ARPA by ACV", "ARPA by MRR", "다중 라인 차트", "22.4Q ~ 25.3Q (12분기)", "억원 / 천만원"],
    ["Row 4", "Revenue on FTE", "Rule of 40", "라인(좌) · 다중 라인(우)", "반기 / 22.4Q ~ 25.3Q", "억원 / %"],
    ["Row 5", "Revenue & Operating Profit", "% Gross Margin & Operating Profit", "콤비네이션 · 다중 라인", "월별 (23.10 ~ 25.10)", "억원 / %"],
    ["Row 6", "Revenue vs. Plan", "Operating Profit vs. Plan", "트리플 바 + 달성률 라인", "월별 (1월 ~ 7월)", "억원 / %"],
]
for i, row_data in enumerate(card_rows):
    r = table_row(ws, r, row_data, even=(i%2==1))

r = sub_title(ws, r, "8-3. 색상 시스템")
r = table_header(ws, r, ["요소", "색상 코드", "", "", "", ""])
color_rows = [
    ["주요 바 차트 (라이트 블루)", "#90CAF9", "", "", "", ""],
    ["미디엄 블루", "#42A5F5", "", "", "", ""],
    ["다크 블루", "#1E88E5", "", "", "", ""],
    ["오렌지 (라인 / YoY)", "#FF9800", "", "", "", ""],
    ["퍼플", "#7E57C2", "", "", "", ""],
    ["그레이", "#9E9E9E", "", "", "", ""],
    ["사이드바 배경", "#1E2A3A ~ #2C3E50 (그라데이션)", "", "", "", ""],
    ["현재 페이지 하이라이트", "#17B9A6", "", "", "", ""],
]
for i, row_data in enumerate(color_rows):
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.cell(row=r, column=1, value=row_data[0]).font = normal_font(size=10)
    ws.cell(row=r, column=2, value=row_data[1]).font = Font(name="Courier New", size=10)
    bg = GRAY_ROW if i%2==1 else WHITE
    for col in [1, 2]:
        ws.cell(row=r, column=col).fill = fill(bg)
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r, column=col).border = border_all("E2E6EE")
    ws.row_dimensions[r].height = 18
    r += 1
r = blank(ws, r)

# ══ 9. 데이터 흐름 ══
r = section_title(ws, r, "9. 데이터 흐름 구조")
flow_lines = [
    "vf219 (매출총계)          ← 당사 DB 기 수신",
    "vf599 (영업/마케팅 비용)  ← 당사 DB 기 수신",
    "      │",
    "      ▼ (지표 산출 로직)",
    "ARPA · 이탈률 · LTV · CAC · LTV:CAC · CAC Payback",
    "      │",
    "      ├─▶ Row 2: CAC / LTV / LTV:CAC Ratio 카드",
    "      ├─▶ Row 2: CAC Payback Period 카드",
    "      │",
    "      ▼ (기존 BI 로직 연동)",
    "Sales Efficiency · ARPA by ACV/MRR · Revenue on FTE · Rule of 40 · Revenue & Operating Profit · Plan 대비",
    "      │",
    "      ▼",
    "효율/수익 대시보드 화면 (Row 1 ~ Row 6)",
]
for line in flow_lines:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    cell = ws.cell(row=r, column=1, value=line)
    cell.font = Font(name="Courier New", size=10, color="1A1A1A")
    cell.fill = fill("F0F4FB")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = border_all("C8D4EC")
    ws.row_dimensions[r].height = 18
    r += 1
r = blank(ws, r)

# ══ 10. 문의처 ══
r = section_title(ws, r, "10. 문의처")
r = table_header(ws, r, ["항목", "내용", "", "", "", ""])
contact_rows = [
    ["담당자", "[담당자명]", "", "", "", ""],
    ["부서", "솔루션사업부 기획팀", "", "", "", ""],
    ["연락처", "[이메일] / [내선번호]", "", "", "", ""],
]
for i, row_data in enumerate(contact_rows):
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.cell(row=r, column=1, value=row_data[0]).font = normal_font(size=10)
    ws.cell(row=r, column=2, value=row_data[1]).font = normal_font(size=10)
    bg = GRAY_ROW if i%2==1 else WHITE
    for col in [1, 2]:
        ws.cell(row=r, column=col).fill = fill(bg)
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r, column=col).border = border_all("E2E6EE")
    ws.row_dimensions[r].height = 18
    r += 1
r = blank(ws, r)

# ══ 푸터 ══
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
cell = ws.cell(row=r, column=1, value="솔루션사업부 기획팀 · 2026-04-20 · v0.2")
cell.font = Font(name="맑은 고딕", size=9, color="9CA3AF")
cell.alignment = Alignment(horizontal="right", vertical="center")
for col in range(1, 7):
    ws.cell(row=r, column=col).border = Border(top=Side(style="thin", color="E0E4EA"))

wb.save("/home/ubuntu/efficiency_profit_dev_request.xlsx")
print("efficiency_profit_dev_request.xlsx 생성 완료")
