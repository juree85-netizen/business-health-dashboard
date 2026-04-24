import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── 스타일 헬퍼 ──────────────────────────────────────────────
H1 = Font(name='맑은 고딕', size=13, bold=True)
H2 = Font(name='맑은 고딕', size=11, bold=True)
BODY = Font(name='맑은 고딕', size=10)
FILL_BLUE = PatternFill("solid", fgColor="1F4E79")
FILL_MID  = PatternFill("solid", fgColor="2E75B6")
FILL_LIGHT= PatternFill("solid", fgColor="D6E4F0")
FILL_GRAY = PatternFill("solid", fgColor="F2F2F2")
WHITE = Font(name='맑은 고딕', size=10, bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val, fill, font=None, align=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill
    c.font = font or WHITE
    c.alignment = align or CENTER
    c.border = BORDER
    return c

def cell(ws, row, col, val, fill=None, font=None, align=None):
    c = ws.cell(row=row, column=col, value=val)
    if fill: c.fill = fill
    c.font = font or BODY
    c.alignment = align or WRAP
    c.border = BORDER
    return c

def title_row(ws, row, text, ncols=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = H2
    c.fill = FILL_LIGHT
    c.alignment = CENTER
    c.border = BORDER

# ══════════════════════════════════════════════════════════════
# Sheet 1: 개요
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1. 요청 개요"
ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 40
ws1.column_dimensions['D'].width = 22
ws1.column_dimensions['E'].width = 22

# 제목
ws1.merge_cells('A1:E1')
c = ws1['A1']
c.value = "VoC 데이터 인터페이스 요청서"
c.font = Font(name='맑은 고딕', size=14, bold=True, color="FFFFFF")
c.fill = FILL_BLUE
c.alignment = CENTER
c.border = BORDER
ws1.row_dimensions[1].height = 32

# 배경
ws1.merge_cells('A3:E3')
title_row(ws1, 3, "0. 배경", 5)
bgs = [
    ("기존 인터페이스", "solution_if_voc 테이블 → /datalake/QualitySummary\nVOC건수·불만건수·실사용자수 3개 지표, 월별 집계(CNT), 1일 1회 배치"),
    ("문제점", "기존 데이터는 월별 합산값 구조 → 개별 건 단위 분석(리드타임·지연 여부) 불가"),
    ("신규 요청 목적", "접수월별 개선건수·평균 리드타임·지연건수 산출을 위해 CS 대시보드로부터 건별 데이터 추가 수신"),
    ("인터페이스 방식", "신규 테이블(건별 데이터) 방식 채택\n기존 테이블(solution_if_voc) 현행 유지\n신규 테이블(solution_if_voc_detail) 추가"),
    ("제공 주기", "CS 대시보드는 스냅샷 방식 운영 → 당사 시계열 트렌드 분석 목적으로 매월 1회 제공 요청"),
]
for i, (k, v) in enumerate(bgs, start=4):
    cell(ws1, i, 1, i-3, FILL_GRAY, Font(name='맑은 고딕', size=10), CENTER)
    cell(ws1, i, 2, k, FILL_LIGHT, Font(name='맑은 고딕', size=10, bold=True), WRAP)
    ws1.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
    cell(ws1, i, 3, v)
    ws1.row_dimensions[i].height = 36

# 테이블 구조
title_row(ws1, 10, "테이블 구조 개요", 5)
hdr(ws1, 11, 1, "#", FILL_MID); hdr(ws1, 11, 2, "테이블", FILL_MID)
ws1.merge_cells(start_row=11, start_column=3, end_row=11, end_column=5)
hdr(ws1, 11, 3, "역할", FILL_MID)
tables = [
    (1, "solution_if_voc", "기존 월별 집계 (현행 유지)"),
    (2, "solution_if_voc_detail", "CS 대시보드로부터 건별 데이터 16개 필드 (신규 인터페이스 요청)"),
    (3, "sol_if_voc_snap", "리드타임·지연건수 월별 스냅샷 — 솔사 내부 개발 예정 (본 인터페이스 범위 아님)"),
]
for i, (n, t, r) in enumerate(tables, start=12):
    cell(ws1, i, 1, n, None, None, CENTER)
    cell(ws1, i, 2, t, FILL_GRAY if n == "solution_if_voc_detail" else None)
    ws1.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
    cell(ws1, i, 3, r)
    ws1.row_dimensions[i].height = 30

# 요청 개요
title_row(ws1, 16, "1. 요청 개요", 5)
hdr(ws1, 17, 1, "#", FILL_MID); hdr(ws1, 17, 2, "항목", FILL_MID)
ws1.merge_cells(start_row=17, start_column=3, end_row=17, end_column=5)
hdr(ws1, 17, 3, "내용", FILL_MID)
items = [
    (1, "요청 목적", "Solution Biz Health Check 대시보드 VoC 지표 산출용 건별 원천 데이터 신규 제공"),
    (2, "활용 시스템", "사업건전성 대시보드 (solution_if_voc_detail 테이블)"),
    (3, "원천 시스템", "CS 대시보드"),
    (4, "기존 인터페이스", "/datalake/QualitySummary (사용자수 / VOC건수 / 불만건수) — 현행 유지"),
]
for i, (n, k, v) in enumerate(items, start=18):
    cell(ws1, i, 1, n, None, None, CENTER)
    cell(ws1, i, 2, k, FILL_LIGHT, Font(name='맑은 고딕', size=10, bold=True))
    ws1.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
    cell(ws1, i, 3, v)
    ws1.row_dimensions[i].height = 28

# ══════════════════════════════════════════════════════════════
# Sheet 2: 요청 필드
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2. 요청 필드")
ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 36
ws2.column_dimensions['D'].width = 28

ws2.merge_cells('A1:D1')
c = ws2['A1']
c.value = "4. 요청 필드 목록 (solution_if_voc_detail)"
c.font = Font(name='맑은 고딕', size=13, bold=True, color="FFFFFF")
c.fill = FILL_BLUE
c.alignment = CENTER
c.border = BORDER
ws2.row_dimensions[1].height = 30

hdr(ws2, 2, 1, "#", FILL_MID)
hdr(ws2, 2, 2, "필드명", FILL_MID)
hdr(ws2, 2, 3, "설명", FILL_MID)
hdr(ws2, 2, 4, "샘플값", FILL_MID)

fields = [
    (1,  "기준월",        "요청일자 기준 월",                                "2026-03"),
    (2,  "VOC ID",       "VoC 고유 식별자",                                 "2603019000025950"),
    (3,  "요청일자",      "VoC 접수 일시",                                   "2026-03-01 00:01:36"),
    (4,  "예정일자",      "처리 예정 기한",                                   "2026-03-03 00:01:36"),
    (5,  "완료일자",      "처리 완료 일시 (미완료건 null)",                   "2026-03-01 00:09:40"),
    (6,  "리드타임(시간)","완료일자 − 요청일자 (시간 단위, 완료건만)",         "0.13"),
    (7,  "진행단계",      "완료/미완료 구분",                                 "완료 / 미완료"),
    (8,  "상품/서비스",   "대상 상품명",                                      "Knox Portal 등"),
    (9,  "처리구분",      "처리 선 구분",                                     "1선 / 2선 / 3선"),
    (10, "접수채널",      "VoC 접수 경로",                                    "CALL / CHAT / E-MAIL / WEB"),
    (11, "요구유형",      "VoC 요구 유형",                                    "개선 / 문의 / 오류 / 기타"),
    (12, "대내외구분",    "고객 대내외 구분",                                  "대내 / 대외"),
    (13, "상담유형(대)",  "상담 유형 대분류",                                  "등록 / 접속 / 사용 등"),
    (14, "상담유형(중)",  "상담 유형 중분류",                                  ""),
    (15, "상담유형(소)",  "상담 유형 소분류",                                  ""),
    (16, "고객사",        "고객사명",                                          "삼성전자 등"),
]
for i, (n, f, d, s) in enumerate(fields, start=3):
    fill = FILL_GRAY if i % 2 == 0 else None
    cell(ws2, i, 1, n, fill, None, CENTER)
    cell(ws2, i, 2, f, fill, Font(name='맑은 고딕', size=10, bold=True))
    cell(ws2, i, 3, d, fill)
    cell(ws2, i, 4, s, fill)
    ws2.row_dimensions[i].height = 22

# ══════════════════════════════════════════════════════════════
# Sheet 3: 제공 조건 & 적재 방식
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3. 제공조건·적재방식")
ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 22
ws3.column_dimensions['C'].width = 44
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 18

ws3.merge_cells('A1:E1')
c = ws3['A1']
c.value = "5. 데이터 제공 조건 (solution_if_voc_detail)"
c.font = Font(name='맑은 고딕', size=13, bold=True, color="FFFFFF")
c.fill = FILL_BLUE
c.alignment = CENTER
c.border = BORDER
ws3.row_dimensions[1].height = 30

hdr(ws3, 2, 1, "#", FILL_MID)
hdr(ws3, 2, 2, "항목", FILL_MID)
ws3.merge_cells(start_row=2, start_column=3, end_row=2, end_column=5)
hdr(ws3, 2, 3, "내용", FILL_MID)

conds = [
    (1, "제공 범위", "2023년 1월 ~ 제공 시점 당월"),
    (2, "제공 주기", "매월 1회"),
    (3, "적재 방식", "3단계 방식 (아래 참조)"),
]
for i, (n, k, v) in enumerate(conds, start=3):
    cell(ws3, i, 1, n, None, None, CENTER)
    cell(ws3, i, 2, k, FILL_LIGHT, Font(name='맑은 고딕', size=10, bold=True))
    ws3.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
    cell(ws3, i, 3, v)
    ws3.row_dimensions[i].height = 24

title_row(ws3, 7, "적재 방식 상세", 5)
hdr(ws3, 8, 1, "단계", FILL_MID)
hdr(ws3, 8, 2, "대상 범위", FILL_MID)
hdr(ws3, 8, 3, "방식", FILL_MID)
hdr(ws3, 8, 4, "시점", FILL_MID)
ws3.merge_cells(start_row=8, start_column=4, end_row=8, end_column=5)

steps = [
    ("① 초기 적재", "2023년 1월 ~ M-2월 전체", "Full 적재", "최초 1회"),
    ("② 변경분 업데이트", "2023년 1월 ~ M-2월 중 변경건", "변경건만 업데이트", "매월"),
    ("③ 전월 적재", "M-1월 전체", "Full 적재", "매월"),
]
for i, (st, rng, way, timing) in enumerate(steps, start=9):
    fill = FILL_GRAY if i % 2 == 0 else None
    cell(ws3, i, 1, st, fill, Font(name='맑은 고딕', size=10, bold=True), CENTER)
    cell(ws3, i, 2, rng, fill)
    cell(ws3, i, 3, way, fill)
    ws3.merge_cells(start_row=i, start_column=4, end_row=i, end_column=5)
    cell(ws3, i, 4, timing, fill, None, CENTER)
    ws3.row_dimensions[i].height = 26

# ══════════════════════════════════════════════════════════════
# Sheet 4: 지표 산출
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4. 산출 지표")
ws4.column_dimensions['A'].width = 5
ws4.column_dimensions['B'].width = 26
ws4.column_dimensions['C'].width = 46
ws4.column_dimensions['D'].width = 30
ws4.column_dimensions['E'].width = 12

ws4.merge_cells('A1:E1')
c = ws4['A1']
c.value = "6. 산출 예정 지표"
c.font = Font(name='맑은 고딕', size=13, bold=True, color="FFFFFF")
c.fill = FILL_BLUE
c.alignment = CENTER
c.border = BORDER
ws4.row_dimensions[1].height = 30

hdr(ws4, 2, 1, "#", FILL_MID)
hdr(ws4, 2, 2, "지표명", FILL_MID)
hdr(ws4, 2, 3, "계산식", FILL_MID)
hdr(ws4, 2, 4, "활용 테이블/필드", FILL_MID)
hdr(ws4, 2, 5, "스냅샷", FILL_MID)

indicators = [
    (1, "접수월별 개선 요청 건수",
     "요구유형 = '개선'인 건수 (기준월별, 상품별)",
     "기준월, 상품/서비스, 요구유형", "✕"),
    (2, "접수월별 평균 처리 리드타임",
     "SUM(리드타임(시간)) ÷ COUNT(진행단계='완료')",
     "기준월, 상품/서비스, 진행단계, 리드타임(시간)", "✓"),
    (3, "월별 지연 건수",
     "완료됐으나 완료일자 > 예정일자\nOR 미완료이며 예정일자 < M월 말일",
     "예정일자, 완료일자, 상품/서비스", "✓"),
]
for i, (n, nm, calc, fld, snap) in enumerate(indicators, start=3):
    fill = FILL_GRAY if i % 2 == 0 else None
    cell(ws4, i, 1, n, fill, None, CENTER)
    cell(ws4, i, 2, nm, fill, Font(name='맑은 고딕', size=10, bold=True))
    cell(ws4, i, 3, calc, fill)
    cell(ws4, i, 4, fld, fill)
    cell(ws4, i, 5, snap, fill, None, CENTER)
    ws4.row_dimensions[i].height = 36

# ══════════════════════════════════════════════════════════════
# Sheet 5: 대상 상품 목록
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5. 대상 상품 목록")
ws5.column_dimensions['A'].width = 5
ws5.column_dimensions['B'].width = 30
ws5.column_dimensions['C'].width = 5
ws5.column_dimensions['D'].width = 30

ws5.merge_cells('A1:D1')
c = ws5['A1']
c.value = "3. 대상 상품/서비스 (32개)"
c.font = Font(name='맑은 고딕', size=13, bold=True, color="FFFFFF")
c.fill = FILL_BLUE
c.alignment = CENTER
c.border = BORDER
ws5.row_dimensions[1].height = 30

hdr(ws5, 2, 1, "#", FILL_MID); hdr(ws5, 2, 2, "상품/서비스", FILL_MID)
hdr(ws5, 2, 3, "#", FILL_MID); hdr(ws5, 2, 4, "상품/서비스", FILL_MID)

products = [
    "Anyframe", "Brightics AI", "Brity Assistant", "Brity Automation",
    "Brity Drive", "Brity Mail", "Brity Mail Mobile", "Brity Meeting",
    "Brity Messenger", "EMM", "EMM (Cloud)", "Endpoint Security",
    "FabriX", "FMC", "IAM", "Knox Drive",
    "Knox EFSS", "Knox Meeting", "Knox Portal", "Knox Portal Mobile",
    "Knox Teams", "Mobile u-Ready", "Nexplant MCS", "Nexplant MES",
    "Nexplant TC", "Nexprime HCM", "Nexprime SCM", "RBS",
    "SCP", "SVPN", "VDI", "인터넷전화",
]
half = len(products) // 2
for i in range(half):
    r = i + 3
    fill = FILL_GRAY if i % 2 == 0 else None
    cell(ws5, r, 1, i+1, fill, None, CENTER)
    cell(ws5, r, 2, products[i], fill)
    cell(ws5, r, 3, i+half+1, fill, None, CENTER)
    cell(ws5, r, 4, products[i+half], fill)
    ws5.row_dimensions[r].height = 22

note_row = half + 3
ws5.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
c = ws5.cell(row=note_row, column=1,
    value="※ CS 대시보드에 신규 상품/서비스 등록 시 별도 요청 없이 자동 반영됩니다.")
c.font = Font(name='맑은 고딕', size=9, italic=True, color="595959")
c.alignment = Alignment(horizontal='left', vertical='center')
ws5.row_dimensions[note_row].height = 20

wb.save('/home/ubuntu/voc_interface_request.xlsx')
print("voc_interface_request.xlsx 생성 완료")
